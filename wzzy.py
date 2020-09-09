# coding=utf-8
import streamlit as st
import numpy as np
import pandas as pd


qzyz = 0.2    #权重因子0-1，就是在一串症状中，有多少比例的症状是有效的

st.title("中医AI问诊V2.0")

import os
path_list = os.listdir()
path_name=[]
for i in path_list:
	if i.split(".")[1] == 'xlsx':
		path_name.append(i.split(".")[0])
zhuanjiaXZ=st.selectbox('请选择一个专家AI:',(path_name))
zhuangjiaXZ_file=str(zhuanjiaXZ)+'.xlsx'
from openpyxl import load_workbook
wb = load_workbook(zhuangjiaXZ_file)
Sheet1 = wb["Sheet1"]
zhenzhuang=set()
st.write(Sheet1.cell(1,1).value)
qzyz = Sheet1.cell(1,12).value     #从表中读取权重因子
if qzyz <0 or qzyz>1 :
	qzyz =0.2
for i in range(3,Sheet1.max_row+1):
	for j in range(1,11):
		zhenzhuang.add(Sheet1.cell(i,j).value)
zhenzhuang.discard(None)
zhenzhuanglist=list(zhenzhuang)
zhenzhuanglist.sort()
zhenzhuangXZ = st.multiselect(
	'选择你的症状（多选）：',(zhenzhuanglist))
zhenzhuangXZ_set=set(zhenzhuangXZ)



#开始症状分析
zhenzhuangGL=set()
yaofang=list()
zhenduan=set()
quanzhong=[0,0,0] #症状对应每个方剂的权重
for i in range(3,Sheet1.max_row+1):
	for j in range(1,11):
		zhenzhuangGL.add(Sheet1.cell(i,j).value)
	zhenzhuangGL.discard(None)
	z = zhenzhuangGL.intersection(zhenzhuangXZ_set)   #两个集合的交集
	if len(zhenzhuangGL)==0 :
		quanzhong.append(len(z))
	else:
		quanzhong.append(len(z)/len(zhenzhuangGL))
	zhenzhuangGL.clear()


import copy 
t = copy.deepcopy(quanzhong)
# 求m个最大的数值及其索引
max_number = list()
max_index = list()
for _ in range(3):
    number = max(t)
    index = t.index(number)
    t[index] = 0
    max_number.append(number)
    max_index.append(index)
t = []

def is_number(s):  #判断是否数字的函数
	s1=['0','1','2','3','4','5','6','7','8','9','+','-'];
	try:
		if s in s1:
			return True
	except ValueError:
		pass
	return False
	
yaofang_name=list()
if max_number[2] < qzyz :  #移除权重过低的答案
	max_number.pop()
	max_index.pop()
if max_number[1] < qzyz :  
	max_number.pop()
	max_index.pop()
if max_number[0] < qzyz :  	
	st.write('未能做出诊断，请重新选择症状')
else:
	for i in range(len(max_index)) :
		zhenduan.add(Sheet1.cell(max_index[i],11).value)   #最接近症状的诊断
		zhenduan.add(Sheet1.cell(max_index[i],12).value)
		for j in range(16,Sheet1.max_column+1):
			yaofang.append(Sheet1.cell(max_index[i],j).value)
	
	yaofang_hanjilian=False   #判断药方里是否含有剂量
	k1=''			
	for k in yaofang:  #k1为药方里面的中药名
		if k:
			k1=k
			if is_number(k[-1]):
				yaofang_hanjilian=True
				k1 = k[:-1]
				if is_number(k[-2]):
					k1 = k[:-2]
					if is_number(k[-3]):
						k1 = k[:-3]
						if is_number(k[-4]):
							k1 = k[:-4]
			if not k1 in yaofang_name:
				yaofang_name.append(k1)  #yaofang_name为纯粹中药名的表
	
	if yaofang_hanjilian:
		yaofang_list=list()
		for l in yaofang_name:   #计算每样药材的量
			l1=0
			for m in yaofang:
				if m:
					l3=''
					if l in m:
						l2=m[len(l):]
						if not l2=='':
							if l2[0]=='+' or l2[0]=='-':
								if l2[0]=='+':
									l1=l1+int(l2[1:])
								if l2[0]=='-':
									l1=l1-int(l2[1:])
							else:
								l1=l1+int(l2)
			if l1>0:
				l3=l + str(l1)
				yaofang_list.append(l3)
	else:
		yaofang_list=yaofang_name
	
	zhenduan.discard(None)
	zhenduan_list=list(zhenduan)
	zhenduan_list.sort()
	str = "，"
	st.write('诊断及建议: ')
	st.write(str.join(zhenduan_list))
	st.write('建议中药：',str.join(yaofang_list))
	st.write('*中药剂量及服用有严格的要求，请在合格中医师指导下进行，我们不对服用中药后引起的任何后果负责')


#© 2020 GitHub, Inc.